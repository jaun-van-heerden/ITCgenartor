from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from Generator_GUI import Ui_MainWindow
import openpyxl as xl
import pandas as pd

from datetime import datetime
import json
import os
import sys
import requests
from urllib.parse import urlencode

files_data = []
files_temp = []
DATA = []

master = {'CurrentWB': '',
          'CurrentWS': '',
          'DATA': {},
          'TEMPLATE': {}}


class MainWindow(QMainWindow, Ui_MainWindow):

    def __init__(self, *args, **kwargs):
        super(MainWindow, self).__init__(*args, **kwargs)
        self.setupUi(self)
        self.TEMPLATES = {}

        # self.templateButton.pressed.connect(self.add_template)
        self.dataButton.clicked.connect(self.on_data_click)
        self.dataCButton.clicked.connect(self.on_dataC_click)
        self.templateButton.clicked.connect(self.on_template_click)
        self.templateCButton.clicked.connect(self.on_templateC_click)
        self.dataTableWidget.cellDoubleClicked.connect(self.on_dataTable_dclick)

        self.refColSpin.valueChanged.connect(self.on_spin)
        self.refRowSpin.valueChanged.connect(self.on_spin)
        # self.sheetDrop.currentIndexChanged.connect(self.sheet_change)   # TODO
        # self.templateTableWidget.event(Qt.RightButton).connect(self.showMenu()) # TODO
        # mouseEvent(Qt.RightButton).connect(self.showMenu())

        self.show()

    def add_template(self):
        print('NICE')
        # print(openFileNamesDialog())

    @pyqtSlot()
    def on_data_click(self):
        options = QFileDialog.Options()
        files, _ = QFileDialog.getOpenFileNames(self, "QFileDialog.getOpenFileNames()", "",
                                                "All Files (*);;Python Files (*.py)", options=options)
        if files:
            self.load_all(files)  # TODO
            self.dataTableWidget.setRowCount(0)
            self.dataTableWidget.setRowCount(len(master['DATA']))
            [self.dataTableWidget.setItem(i - 1, 1, QTableWidgetItem(v)) for i, v in enumerate(master['DATA'])]

    @pyqtSlot()
    def on_dataC_click(self):
        master['CurrentWS'] = ''
        master['CurrentWB'] = ''
        master['DATA'] = {}
        self.dataTableWidget.setRowCount(0)
        self.tableWidget2.setRowCount(0)
        self.tableWidget2.setColumnCount(0)

    @pyqtSlot()
    def on_template_click(self):
        options = QFileDialog.Options()
        files, _ = QFileDialog.getOpenFileNames(self, "QFileDialog.getOpenFileNames()", "",
                                                "All Files (*);;Python Files (*.py)", options=options)
        if files:
            self.TEMPLATES = self.load_all(self.TEMPLATES, files)  # TODO
            # files_temp = [files_temp.append(i) for i in files if i not in self.TEMPLATES.items()]
            self.templateTableWidget.setRowCount(0)
            self.templateTableWidget.setRowCount(len(self.TEMPLATES))
            [self.templateTableWidget.setItem(i - 1, 1, QTableWidgetItem(v)) for i, v in enumerate(self.TEMPLATES)]
            # self.load_all_csv(self.TEMPLATES, files)
            print(self.TEMPLATES)

    @pyqtSlot()
    def on_templateC_click(self):
        master['TEMPLATE'] = {}
        self.templateTableWidget.setRowCount(0)

    @pyqtSlot()
    def on_dataTable_dclick(self):
        print(self.dataTableWidget.currentItem().text())
        master['CurrentWB'] = self.dataTableWidget.currentItem().text()
        master['CurrentWS'] = ''
        self.display_data(self.dataTableWidget.currentItem().text())
        # self.dataTableWidget.c
        # self.TEMPLATES = {}
        # self.templateTableWidget.setRowCount(0)

    # @pyqtSlot()   # TODO - Sheet change with dropdown
    # def sheet_change(self):
    # self.display_data(self.dataTableWidget.currentItem().text(), self.sheetDrop.currentText())

    @pyqtSlot()
    def on_spin(self):
        try:
            # print(self.refColSpin.value())
            # self.tableWidget2.setItem(3, int(self.refColSpin.value(), QTableWidgetItem()))
            # self.tableWidget2.item(3, int(self.refColSpin.value())).setBackground(QColor(100, 100, 100))
            # self.tableWidget2.setItem(1, 1, QTableWidgetItem())

            for i in range(self.tableWidget2.rowCount()):
                for j in range(self.tableWidget2.columnCount()):
                    self.tableWidget2.item(i, j).setBackground(QColor(Qt.white))

            # [self.tableWidget2.item(i, int(self.refColSpin.value())).setBackground(QColor(Qt.blue)) for i in
            # range(self.tableWidget2.columnCount())]

            try:
                [self.tableWidget2.item(i, int(self.refColSpin.value() - 1)).setBackground(QColor(121, 252, 50, 20)) for
                 i in
                 range(self.tableWidget2.columnCount() + 1)]
            except:
                pass

            try:
                [self.tableWidget2.item(int(self.refRowSpin.value() - 1), i).setBackground(QColor(121, 252, 50, 20)) for
                 i
                 in range(self.tableWidget2.rowCount() + 1)]
            except:
                pass
            # self.tableWidget2.item(1, int(self.refColSpin.value())).setBackground(QColor(200, 200, 200))
            # self.tableWidget2.item(1, int(self.refColSpin.value())-1).setBackground(QColor(110, 0, 10))
            # self.tableWidget2.item(1, int(self.refColSpin.value())).setBackground(QColor(200, 200, 200))
        except:
            print('no')

    # def showMenu(self, event): # TODO
    # print("hello")
    # menu = QMenu()
    # clear_action = menu.addAction("Clear Selection", self)
    # action = menu.exec_(self.mapToGlobal(event.pos()))
    # if action == clear_action:
    # self.clearSelection()

    def contextMenuEvent(self, event):  # TODO
        contextMenu = QMenu(self)
        newAct = contextMenu.addAction("New")
        openAct = contextMenu.addAction("Open")
        quitAct = contextMenu.addAction("Quit")
        action = contextMenu.exec_(self.mapToGlobal(event.pos()))
        if action == quitAct:
            self.close()

    def load_all(self, files):
        for i in files:
            if i.split('/')[-1] not in master['DATA'].items():
                try:
                    wb = xl.load_workbook(filename=i, read_only=True)
                    master['DATA'][i.split('/')[-1].replace('.' + i.split('.')[-1], '')] = [wb, wb.sheetnames]

                    if master['CurrentWB'] == '':
                        master['CurrentWB'] = wb
                        master['CurrentWS'] = wb.sheetnames[0]
                except:
                    print('! Could not load')
        print(master)




    def display_data(self, ref, sheet=None):
        # print(self.DATA)
        # print(self.DATA[ref].sheetnames)
        # Get Max Col
        # print(self.DATA[ref]['A1'])

        try:
            self.sheetDrop.clear()
            self.sheetDrop.addItems(master['DATA'][ref][1])
        except:
            print('NOPE')

        try:
            if sheet is None:
                master['DATA'][ref].active = 1
                ws = master['DATA'][ref]  # = 1 #[master['DATA'][ref][1][0]]
                print(ws)
                   # TODO - Fill the combo box with sheet selection (consider multiple sheet data)


            ws_r_max, ws_c_max = ws.max_row, ws.max_column
            self.tableWidget2.setRowCount(0)
            self.tableWidget2.setColumnCount(0)
            self.tableWidget2.setRowCount(ws_r_max)
            self.tableWidget2.setColumnCount(ws_c_max)

            # for row in ws.iter_rows(values_only=True):
            # for value in row:
            # df = pd.DataFrame(ws.values)    # TODO - Pandas Dataframe
            # print(df.head(50))

            for i in range(ws_r_max):
                for j in range(ws_c_max):
                    # if i % 50 == 0 and j % 50 == 0:    # TODO - Limit the blank columns and rows
                    # if any([i == None for i in range()])

                    try:
                        self.tableWidget2.setItem(i, j, QTableWidgetItem(ws.cell(column=j + 1, row=i + 1).value))
                    except:
                        print('not work at r/c', i, j)
        except:
            print('no')

        # print(ws.calculate_dimension())    # TODO
        # print(ws.reset_dimensions())


if __name__ == '__main__':
    app = QApplication([])
    window = MainWindow()
    app.exec_()

    # from concurrent.futures import ProcessPoolExecutor
    # from openpyxl import load_workbook
    # from time import perf_counter

    # test_file = "Issues/bug494.xlsx"

    # def parallel_worksheet(sheetname):
    # begin = perf_counter()
    # wb = load_workbook(test_file, read_only=True,
    #                        data_only=True, keep_links=False)
    #     ws = wb[sheetname]
    #     for row in ws.iter_rows(values_only=True):
    #         for value in row:
    #             pass
    #     end = perf_counter()
    #     print("    {0} {1:.2f}s".format(sheetname, end - begin))
    #
    #
    # def parallel_read():
    #     print("Parallised Read")
    #     begin = perf_counter()
    #     wb = load_workbook(test_file, read_only=True,
    #                        keep_links=False, data_only=True)
    #     print("    Workbook loaded {0:.2f}s".format(perf_counter() - begin))
    #     sheets = wb.sheetnames
    #     with ProcessPoolExecutor() as pool:
    #         for ws in pool.map(parallel_worksheet, sheets):
    #             pass
    #     end = perf_counter()
    #     print("    Total time {0:.2f}s".format(end - begin))
    #
    #
    # if __name__ == "__main__":
    #     parallel_read()
